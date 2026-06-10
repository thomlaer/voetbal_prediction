"""Genereer een opgemaakte Excel met alle WK 2026 wedstrijdvoorspellingen.

Produceert: outputs_worldcup2026_default/WK2026_Voorspellingen.xlsx
    - Groepsfase + knockout (teams al ingevuld op basis van bracket)
    - Kolommen: teams | score | ODDS INTERNET | ALLEEN MODEL
    - Kleurcodering per ronde, dikgedrukte voorspelde winnaar
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("outputs_worldcup2026_default")
SCHEDULE_PATH = Path("data/extracted/oddsportal_worldcup2026_fixture_odds_schedule.csv")

# ---------------------------------------------------------------------------
# Kleuren
# ---------------------------------------------------------------------------
CLR_TITLE_BG   = "1F4E79"
CLR_TITLE_FG   = "FFFFFF"
CLR_SEC_MATCH  = "BDD7EE"   # licht blauw – wedstrijdinformatie header
CLR_SEC_SCORE  = "D9E1F2"   # licht paars – score header
CLR_SEC_ODDS   = "2F75B6"   # donker blauw – odds internet header
CLR_SEC_MODEL  = "375623"   # donker groen – alleen model header
CLR_COL_HDR    = "F2F2F2"   # licht grijs – kolomkoppen
CLR_GRP_ROW    = "FFFFFF"   # wit – groepsfase rijen
CLR_R32_ROW    = "EBF3FB"   # zeer licht blauw
CLR_R16_ROW    = "FFF2CC"   # licht geel
CLR_QF_ROW     = "FCE4D6"   # licht oranje
CLR_SF_ROW     = "FFD7D7"   # licht rood
CLR_3RD_ROW    = "E2EFDA"   # licht groen
CLR_FIN_ROW    = "FFF2CC"   # goud-tint
CLR_FIN_BG     = "FFD700"   # goud finale cel


STAGE_CLR = {
    "Groepsfase":            CLR_GRP_ROW,
    "Ronde van 32":          CLR_R32_ROW,
    "Ronde van 16":          CLR_R16_ROW,
    "Kwartfinales":          CLR_QF_ROW,
    "Halve finales":         CLR_SF_ROW,
    "3e Plaatswedstrijd":    CLR_3RD_ROW,
    "Finale":                CLR_FIN_ROW,
}

STAGE_NL = {
    "Group Stage":           "Groepsfase",
    "Round of 32":           "Ronde van 32",
    "Round of 16":           "Ronde van 16",
    "Quarterfinals":         "Kwartfinales",
    "Semifinals":            "Halve finales",
    "Third Place Playoff":   "3e Plaatswedstrijd",
    "Final":                 "Finale",
}


# ---------------------------------------------------------------------------
# Score schatting knockout
# ---------------------------------------------------------------------------
def estimate_ko_score(winner: str, home: str, p_win: float) -> str:
    if p_win >= 0.82:
        base = "2-0"
    elif p_win >= 0.65:
        base = "1-0"
    else:
        base = "2-1"
    if winner == home:
        return base
    h, a = base.split("-")
    return f"{a}-{h}"


# ---------------------------------------------------------------------------
# Data laden en samenvoegen
# ---------------------------------------------------------------------------
def load_data(output_dir: Path, schedule_path: Path) -> pd.DataFrame:
    group = pd.read_csv(output_dir / "worldcup2026_group_match_predictions.csv")
    bracket = pd.read_csv(output_dir / "worldcup2026_bracket_prediction.csv")
    schedule = pd.read_csv(schedule_path, usecols=["match_number", "date", "stage"])
    schedule["date"] = pd.to_datetime(schedule["date"], errors="coerce").dt.strftime("%d-%m-%Y")

    # --- groepsfase ---
    group["date"] = pd.to_datetime(group["date"], errors="coerce").dt.strftime("%d-%m-%Y")
    group["stage_nl"] = "Groepsfase"
    group["score_model"] = group["pool_predicted_score"]
    group["score_sim"] = group["sim_predicted_score"]
    group["odds_h"] = group["home_odds"].round(2)
    group["odds_g"] = group["draw_odds"].round(2)
    group["odds_u"] = group["away_odds"].round(2)
    group["pct_odds_h"] = (group["bookie_prob_home_win"] * 100).round(1)
    group["pct_odds_g"] = (group["bookie_prob_draw"] * 100).round(1)
    group["pct_odds_u"] = (group["bookie_prob_away_win"] * 100).round(1)
    group["pct_model_h"] = (group["prob_home_win"] * 100).round(1)
    group["pct_model_g"] = (group["prob_draw"] * 100).round(1)
    group["pct_model_u"] = (group["prob_away_win"] * 100).round(1)
    group["sim_winner"] = group.apply(
        lambda r: r["home_team"] if r["sim_predicted_outcome"] == "home_win"
        else (r["away_team"] if r["sim_predicted_outcome"] == "away_win" else "gelijkspel"),
        axis=1,
    )
    group["groep"] = group["group"]

    group_rows = group[[
        "match_number", "date", "stage_nl", "groep",
        "home_team", "away_team", "city",
        "score_model", "score_sim", "sim_winner",
        "odds_h", "odds_g", "odds_u",
        "pct_odds_h", "pct_odds_g", "pct_odds_u",
        "pct_model_h", "pct_model_g", "pct_model_u",
    ]].copy()

    # --- knockout ---
    bracket = bracket.merge(
        schedule.rename(columns={"stage": "stage_en"}), on="match_number", how="left"
    )
    bracket["stage_nl"] = bracket["stage"].map(STAGE_NL).fillna(bracket["stage"])
    bracket["groep"] = ""
    bracket["score_model"] = bracket.apply(
        lambda r: estimate_ko_score(r["predicted_winner"], r["home_team"], r["winner_win_prob"]),
        axis=1,
    )
    bracket["score_sim"] = bracket["score_model"]
    bracket["sim_winner"] = bracket["predicted_winner"]
    bracket["odds_h"] = None
    bracket["odds_g"] = None
    bracket["odds_u"] = None
    bracket["pct_odds_h"] = None
    bracket["pct_odds_g"] = None
    bracket["pct_odds_u"] = None
    bracket["pct_model_h"] = bracket.apply(
        lambda r: round(r["winner_win_prob"] * 100, 1) if r["predicted_winner"] == r["home_team"]
        else round((1 - r["winner_win_prob"]) * 100, 1),
        axis=1,
    )
    bracket["pct_model_g"] = None
    bracket["pct_model_u"] = bracket.apply(
        lambda r: round(r["winner_win_prob"] * 100, 1) if r["predicted_winner"] == r["away_team"]
        else round((1 - r["winner_win_prob"]) * 100, 1),
        axis=1,
    )

    ko_rows = bracket[[
        "match_number", "date", "stage_nl", "groep",
        "home_team", "away_team", "city",
        "score_model", "score_sim", "sim_winner",
        "odds_h", "odds_g", "odds_u",
        "pct_odds_h", "pct_odds_g", "pct_odds_u",
        "pct_model_h", "pct_model_g", "pct_model_u",
    ]].copy()

    combined = pd.concat([group_rows, ko_rows], ignore_index=True).sort_values("match_number")
    return combined


# ---------------------------------------------------------------------------
# Stijlhelpers
# ---------------------------------------------------------------------------
def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)


def thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def set_cell(ws, row: int, col: int, value, bold=False, color="000000",
             bg=None, align="left", number_format=None, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = fill(bg)
    if number_format:
        cell.number_format = number_format
    cell.border = thin_border()
    return cell


# ---------------------------------------------------------------------------
# Excel bouwen
# ---------------------------------------------------------------------------
def build_excel(df: pd.DataFrame, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "WK 2026 Voorspellingen"

    COLS = {
        "nr":          1,   # A
        "datum":       2,   # B
        "ronde":       3,   # C
        "groep":       4,   # D
        "thuis":       5,   # E
        "uit":         6,   # F
        "stad":        7,   # G
        "sc_model":    8,   # H   Score (model)
        "sc_sim":      9,   # I   Score (odds+model)
        "odds_h":      10,  # J
        "odds_g":      11,  # K
        "odds_u":      12,  # L
        "p_odds_h":    13,  # M
        "p_odds_g":    14,  # N
        "p_odds_u":    15,  # O
        "p_mod_h":     16,  # P
        "p_mod_g":     17,  # Q
        "p_mod_u":     18,  # R
    }
    NCOLS = 18

    # --- rij 1: titel ---
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    c = ws.cell(row=1, column=1, value="WK 2026 - VOORSPELLINGEN ALLE WEDSTRIJDEN")
    c.font = Font(name="Arial", size=14, bold=True, color=CLR_TITLE_FG)
    c.fill = fill(CLR_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- rij 2: sectiekoppen (merged) ---
    ws.row_dimensions[2].height = 22
    sections = [
        (1, 7,  "WEDSTRIJD INFORMATIE",     CLR_SEC_MATCH, "FFFFFF", False),
        (8, 9,  "SCORE VOORSPELLING",        CLR_SEC_SCORE, "000000", False),
        (10, 15, "ODDS INTERNET",             CLR_SEC_ODDS,  "FFFFFF", True),
        (16, 18, "ALLEEN MODEL",              CLR_SEC_MODEL, "FFFFFF", True),
    ]
    for c1, c2, label, bg, fg, bold in sections:
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        cell = ws.cell(row=2, column=c1, value=label)
        cell.font = Font(name="Arial", size=10, bold=True, color=fg)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- rij 3: kolomkoppen ---
    ws.row_dimensions[3].height = 40
    headers = [
        "Nr", "Datum", "Ronde", "Gr.", "Thuisploeg", "Uitploeg", "Stad",
        "Score\n(model)", "Score\n(sim)",
        "Odds\nThuis", "Odds\nGelijk", "Odds\nUit",
        "% Thuis\n(odds)", "% Gelijk\n(odds)", "% Uit\n(odds)",
        "% Thuis\n(model)", "% Gelijk\n(model)", "% Uit\n(model)",
    ]
    for col, hdr in enumerate(headers, 1):
        set_cell(ws, 3, col, hdr, bold=True, bg=CLR_COL_HDR, align="center", wrap=True)

    # --- kolombreedtes ---
    col_widths = {
        1: 5,   # Nr
        2: 12,  # Datum
        3: 18,  # Ronde
        4: 5,   # Groep
        5: 22,  # Thuisploeg
        6: 22,  # Uitploeg
        7: 20,  # Stad
        8: 10,  # Score model
        9: 10,  # Score sim
        10: 8,  # Odds H
        11: 8,  # Odds G
        12: 8,  # Odds U
        13: 9,  # % H odds
        14: 9,  # % G odds
        15: 9,  # % U odds
        16: 9,  # % H model
        17: 9,  # % G model
        18: 9,  # % U model
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # --- data rijen ---
    def pct_str(val) -> str | None:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return "-"
        return f"{val:.1f}%"

    def odds_str(val) -> str | None:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return "-"
        return f"{val:.2f}"

    prev_stage = None
    for data_row_idx, row in enumerate(df.itertuples(index=False), start=4):
        ws.row_dimensions[data_row_idx].height = 20
        stage = row.stage_nl
        bg = STAGE_CLR.get(stage, CLR_GRP_ROW)
        is_final = stage == "Finale"

        # stage-scheidingslijn: extra dikke top rand bij nieuwe ronde
        if stage != prev_stage and prev_stage is not None:
            for col in range(1, NCOLS + 1):
                cell = ws.cell(row=data_row_idx, column=col)
                thick = Side(style="medium", color="1F4E79")
                thin_s = Side(style="thin", color="BFBFBF")
                cell.border = Border(
                    top=thick, bottom=thin_s, left=thin_s, right=thin_s
                )
        prev_stage = stage

        def put(col_key, val, bold=False, align="left", is_score=False):
            col = COLS[col_key]
            extra_bg = CLR_FIN_BG if is_final and is_score else bg
            cell = set_cell(ws, data_row_idx, col, val,
                            bold=bold or is_final,
                            bg=extra_bg, align=align)
            return cell

        put("nr",     row.match_number, align="center")
        put("datum",  row.date, align="center")
        put("ronde",  stage)
        put("groep",  row.groep, align="center")

        # thuisploeg: dikgedrukt als voorspelde winnaar
        thuis_bold = (row.sim_winner == row.home_team)
        uit_bold   = (row.sim_winner == row.away_team)
        set_cell(ws, data_row_idx, COLS["thuis"], row.home_team,
                 bold=thuis_bold or is_final, bg=CLR_FIN_BG if is_final else bg)
        set_cell(ws, data_row_idx, COLS["uit"], row.away_team,
                 bold=uit_bold or is_final, bg=CLR_FIN_BG if is_final else bg)

        put("stad",    row.city)
        put("sc_model", row.score_model, align="center", bold=True, is_score=True)
        put("sc_sim",   row.score_sim,   align="center", bold=True, is_score=True)

        put("odds_h", odds_str(row.odds_h), align="center")
        put("odds_g", odds_str(row.odds_g), align="center")
        put("odds_u", odds_str(row.odds_u), align="center")

        put("p_odds_h", pct_str(row.pct_odds_h), align="center")
        put("p_odds_g", pct_str(row.pct_odds_g), align="center")
        put("p_odds_u", pct_str(row.pct_odds_u), align="center")

        put("p_mod_h", pct_str(row.pct_model_h), align="center")
        put("p_mod_g", pct_str(row.pct_model_g), align="center")
        put("p_mod_u", pct_str(row.pct_model_u), align="center")

    # --- stage-kleurlegenda onderaan ---
    legend_row = len(df) + 6
    ws.cell(row=legend_row, column=1, value="LEGENDA RONDES").font = Font(
        name="Arial", size=10, bold=True
    )
    for i, (stage_name, clr) in enumerate(STAGE_CLR.items(), start=1):
        r = legend_row + i
        c = ws.cell(row=r, column=1, value=stage_name)
        c.fill = fill(clr)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")

    # --- bevroren rijen en kolommen (vastgezet bij rij 4, kolom E) ---
    ws.freeze_panes = "E4"

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Genereer een opgemaakte WK 2026 Excel.")
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--schedule", type=Path, default=SCHEDULE_PATH)
    args = parser.parse_args()

    print("Data laden...")
    df = load_data(args.output_dir, args.schedule)
    out_path = args.output_dir / "WK2026_Voorspellingen.xlsx"
    print(f"Excel aanmaken: {out_path}")
    build_excel(df, out_path)
    print(f"Klaar! {len(df)} wedstrijden verwerkt.")
    print(f"Bestand: {out_path.resolve()}")


if __name__ == "__main__":
    main()
